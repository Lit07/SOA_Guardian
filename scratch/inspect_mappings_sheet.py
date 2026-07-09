import openpyxl

wb = openpyxl.load_workbook("uploads/map_7c5a0799.xlsx")
print("Sheets:", wb.sheetnames)
sheet = wb["Final_mappings"]
print("Header row:")
header = [cell.value for cell in sheet[1]]
print("  ", header)

print("Rows:")
for r_idx in range(2, min(20, sheet.max_row + 1)):
    row = [cell.value for cell in sheet[r_idx]]
    print(f"  Row {r_idx}: {row}")
