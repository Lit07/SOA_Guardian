import os
import openpyxl
import pytest
from soa_guardian.models import CanonicalStatement, StatementMetadata, Transaction
from soa_guardian.exporter import export_to_excel

def create_dummy_template(file_path: str):
    """Creates a dummy target template Excel sheet with a custom set and order of headers."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Template Sheet"
    
    # Non-standard headers order
    headers = [
        "Date", 
        "Particulars (Description)", 
        "Debit", 
        "Credit", 
        "Cum Balance", 
        "Universal Dr (SGD)", 
        "Universal Cr (SGD)", 
        "Universal Bal (SGD)", 
        "Status", 
        "Manual Review Notes"
    ]
    ws.append(headers)
    wb.save(file_path)

def test_template_export_dynamic_mapping():
    """Verify that exporter copies template and maps canonical transaction fields to correct columns."""
    template_path = "tests/dummy_template.xlsx"
    export_path = "tests/dummy_templated_export.xlsx"
    
    create_dummy_template(template_path)
    
    # Instantiate dummy statement
    metadata = StatementMetadata(
        bank_name="Test Bank",
        account_number="ACC-123",
        original_currency="USD",
        universal_currency="SGD",
        exchange_rate=1.35
    )
    
    txs = [
        Transaction(
            transaction_date="01/01/2026",
            description="Salary Payment",
            debit_amount=None,
            credit_amount=5000.00,
            running_balance=5000.00,
            status="clean",
            confidence=1.0,
            universal_debit=None,
            universal_credit=6750.00,
            universal_balance=6750.00
        ),
        Transaction(
            transaction_date="02/01/2026",
            description="Office Supplies",
            debit_amount=150.00,
            credit_amount=None,
            running_balance=4850.00,
            status="auto_repaired",
            confidence=0.9,
            universal_debit=202.50,
            universal_credit=None,
            universal_balance=6547.50
        )
    ]
    
    canonical = CanonicalStatement(
        statement_metadata=metadata,
        opening_balance=0.0,
        closing_balance=4850.00,
        transactions=txs,
        confidence=0.95,
        extraction_method="pdf",
        source_page="1",
        anomaly_flags=[]
    )
    
    try:
        # Export to template
        export_to_excel(
            canonical=canonical,
            output_path=export_path,
            flat_table=False,
            template_path=template_path
        )
        
        # Verify the exported workbook structure
        assert os.path.exists(export_path)
        wb = openpyxl.load_workbook(export_path, data_only=True)
        ws = wb.active
        
        # Verify headers are exactly from template
        assert ws.cell(row=1, column=1).value == "Date"
        assert ws.cell(row=1, column=2).value == "Particulars (Description)"
        assert ws.cell(row=1, column=3).value == "Debit"
        assert ws.cell(row=1, column=4).value == "Credit"
        assert ws.cell(row=1, column=5).value == "Cum Balance"
        assert ws.cell(row=1, column=6).value == "Universal Dr (SGD)"
        assert ws.cell(row=1, column=7).value == "Universal Cr (SGD)"
        assert ws.cell(row=1, column=8).value == "Universal Bal (SGD)"
        assert ws.cell(row=1, column=9).value == "Status"
        
        # Verify Row 2 values (Salary Payment)
        # "Date" -> "01/01/2026"
        assert ws.cell(row=2, column=1).value == "01/01/2026"
        # "Particulars (Description)" -> "Salary Payment"
        assert ws.cell(row=2, column=2).value == "Salary Payment"
        # "Debit" -> empty
        assert ws.cell(row=2, column=3).value is None
        # "Credit" -> 5000.00
        assert ws.cell(row=2, column=4).value == 5000.00
        # "Cum Balance" -> 5000.00
        assert ws.cell(row=2, column=5).value == 5000.00
        # "Universal Dr (SGD)" -> empty
        assert ws.cell(row=2, column=6).value is None
        # "Universal Cr (SGD)" -> 6750.00
        assert ws.cell(row=2, column=7).value == 6750.00
        # "Universal Bal (SGD)" -> 6750.00
        assert ws.cell(row=2, column=8).value == 6750.00
        # "Status" -> "clean"
        assert ws.cell(row=2, column=9).value == "clean"
        
        # Verify Row 3 values (Office Supplies)
        assert ws.cell(row=3, column=1).value == "02/01/2026"
        assert ws.cell(row=3, column=2).value == "Office Supplies"
        assert ws.cell(row=3, column=3).value == 150.00
        assert ws.cell(row=3, column=4).value is None
        assert ws.cell(row=3, column=5).value == 4850.00
        assert ws.cell(row=3, column=6).value == 202.50
        assert ws.cell(row=3, column=7).value is None
        assert ws.cell(row=3, column=8).value == 6547.50
        assert ws.cell(row=3, column=9).value == "auto_repaired"
        
    finally:
        if os.path.exists(template_path):
            os.remove(template_path)
        if os.path.exists(export_path):
            os.remove(export_path)
