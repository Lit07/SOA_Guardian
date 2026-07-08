import os
import csv
import re
import pytest
import pandas as pd
import openpyxl
from soa_guardian.scrubber import scrub_file

def create_raw_test_csv(file_path: str):
    """Creates a raw synthetic CSV statement of account containing PII."""
    data = [
        ["Contact Person: Jane Doe (jane.doe@corporation.com)", "", "", "", ""],
        ["Phone Support: +91 98765 43210", "", "", "", ""],
        ["Vendor Account No: 88776655", "", "", "", ""],
        ["Opening Balance: 2000.00", "", "", "", ""],
        [],
        ["Date", "Description", "Withdrawals (Dr)", "Deposits (Cr)", "Balance"],
        ["01/06/2026", "Opening Balance", "", "", "2000.00"],
        ["02/06/2026", "Purchase invoice #A1", "350.00", "", "1650.00"],
        ["03/06/2026", "Client service payment", "", "1200.00", "2850.00"],
        [],
        ["Closing Balance: 2850.00", "", "", "", ""]
    ]
    with open(file_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(data)

def test_csv_scrubbing_text_only():
    """Verify that text-only scrubbing anonymizes names/PII but keeps numeric values unchanged."""
    input_csv = "tests/raw_input.csv"
    output_csv = "tests/scrubbed_output.csv"
    create_raw_test_csv(input_csv)
    
    try:
        # Scrub without randomizing amounts
        scrub_file(input_csv, output_csv, randomize_amounts=False)
        assert os.path.exists(output_csv)
        
        with open(output_csv, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            rows = list(reader)
            
        # Verify metadata is scrubbed
        assert "jane.doe" not in rows[0][0]
        assert "contact@mockvendor.com" in rows[0][0]
        assert "+91 98765" not in rows[1][0]
        assert "+1-555-0199" in rows[1][0]
        assert "88776655" not in rows[2][0]
        assert "ACC-XXXXXX" in rows[2][0]
        
        # Verify transaction description is scrubbed
        assert "Purchase invoice #A1" not in rows[7][1]
        assert any(term in rows[7][1] for term in ["Batch", "Invoice", "Procurement", "Fee", "Retainer"])
        
        # Verify numeric amounts are unchanged
        assert rows[7][2] == "350.00"
        assert rows[8][3] == "1200.00"
        assert rows[8][4] == "2850.00"
        assert "2850.00" in rows[10][0]
        
    finally:
        if os.path.exists(input_csv):
            os.remove(input_csv)
        if os.path.exists(output_csv):
            os.remove(output_csv)

def test_xlsx_scrubbing_randomized_amounts():
    """Verify that randomized amounts scrubbing scrambles figures and keeps running balances mathematically consistent."""
    input_xlsx = "tests/raw_input.xlsx"
    output_xlsx = "tests/scrubbed_output.xlsx"
    
    # Generate XLSX file matching structure directly using openpyxl
    wb_temp = openpyxl.Workbook()
    ws_temp = wb_temp.active
    data = [
        ["Contact Person: Jane Doe (jane.doe@corporation.com)", "", "", "", ""],
        ["Phone Support: +91 98765 43210", "", "", "", ""],
        ["Vendor Account No: 88776655", "", "", "", ""],
        ["Opening Balance: 2000.00", "", "", "", ""],
        [],
        ["Date", "Description", "Withdrawals (Dr)", "Deposits (Cr)", "Balance"],
        ["01/06/2026", "Opening Balance", "", "", 2000.00],
        ["02/06/2026", "Purchase invoice #A1", 350.00, "", 1650.00],
        ["03/06/2026", "Client service payment", "", 1200.00, 2850.00],
        [],
        ["Closing Balance: 2850.00", "", "", "", ""]
    ]
    for r in data:
        ws_temp.append(r)
    wb_temp.save(input_xlsx)
    
    try:
        # Scrub with amount randomization enabled
        scrub_file(input_xlsx, output_xlsx, randomize_amounts=True)
        assert os.path.exists(output_xlsx)
        
        wb = openpyxl.load_workbook(output_xlsx, data_only=True)
        ws = wb.active
        
        # Verify metadata is scrubbed
        r1_val = ws.cell(row=1, column=1).value
        r2_val = ws.cell(row=2, column=1).value
        r3_val = ws.cell(row=3, column=1).value
        
        assert "jane.doe" not in r1_val
        assert "contact@mockvendor.com" in r1_val
        assert "+1-555-0199" in r2_val
        assert "ACC-XXXXXX" in r3_val
        
        # Extract randomized mathematical elements
        op_bal_cell = ws.cell(row=4, column=1).value
        match_op = re.search(r'([\d\.]+)', op_bal_cell)
        op_bal = float(match_op.group(1))
        
        # First row: opening balance
        t1_bal = float(ws.cell(row=7, column=5).value)
        assert t1_bal == op_bal
        
        # Second row: debit transaction (originally 350.00)
        t2_deb = float(ws.cell(row=8, column=3).value or 0.0)
        t2_cred = float(ws.cell(row=8, column=4).value or 0.0)
        t2_bal = float(ws.cell(row=8, column=5).value)
        
        assert t2_deb > 0.0
        assert t2_deb != 350.00  # Verify it was randomized
        assert t2_cred == 0.0
        # Check cumulative sum balance math is consistent
        assert abs(t2_bal - (t1_bal + t2_cred - t2_deb)) <= 0.01
        
        # Third row: credit transaction (originally 1200.00)
        t3_deb = float(ws.cell(row=9, column=3).value or 0.0)
        t3_cred = float(ws.cell(row=9, column=4).value or 0.0)
        t3_bal = float(ws.cell(row=9, column=5).value)
        
        assert t3_deb == 0.0
        assert t3_cred > 0.0
        assert t3_cred != 1200.00  # Verify it was randomized
        # Check cumulative sum balance math is consistent
        assert abs(t3_bal - (t2_bal + t3_cred - t3_deb)) <= 0.01
        
        # Closing balance block check
        footer_val = ws.cell(row=11, column=1).value
        match_cl = re.search(r'([\d\.]+)', footer_val)
        cl_bal = float(match_cl.group(1))
        # Ensure closing balance was rewritten to equal the final running balance cell
        assert abs(cl_bal - t3_bal) <= 0.01
        
    finally:
        if os.path.exists(input_xlsx):
            os.remove(input_xlsx)
        if os.path.exists(output_xlsx):
            os.remove(output_xlsx)
