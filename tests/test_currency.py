import os
import openpyxl
import pytest
from soa_guardian.validator import parse_currency
from soa_guardian.pipeline import process_statement
from tests.generate_test_data import create_clean_pdf, create_clean_xlsx

def test_parenthetical_negative_parsing():
    """Verify that parse_currency correctly identifies parenthetical values as negative numbers."""
    assert parse_currency("(1,020.98)", "period_decimal") == -1020.98
    assert parse_currency("($120.50)", "period_decimal") == -120.50
    assert parse_currency("( 50.00 )", "period_decimal") == -50.00
    assert parse_currency("(1.020,98)", "comma_decimal") == -1020.98
    assert parse_currency("(-100.00)", "period_decimal") == -100.00 # absolute checks avoid double negation

def test_pipeline_currency_conversion():
    """Verify that process_statement maps currency metadata and calculates universal columns correctly."""
    pdf_path = "tests/clean_pdf_statement.pdf"
    assert os.path.exists(pdf_path)
    
    # Process with exchange rate (e.g., USD to EUR rate = 0.74)
    canonical = process_statement(
        pdf_path, 
        universal_currency="EUR",
        exchange_rate=0.74
    )
    
    # Assert metadata currency settings
    assert canonical.statement_metadata.original_currency == "USD" # Default fallback when no currency cues present
    assert canonical.statement_metadata.universal_currency == "EUR"
    assert canonical.statement_metadata.exchange_rate == 0.74
    
    # Verify transaction calculations (Row 2: Salary Credit 2500.00 * 0.74 = 1850.00)
    tx = canonical.transactions[1]
    assert tx.credit_amount == 2500.00
    assert tx.universal_credit == 1850.00
    assert tx.universal_balance == round(tx.running_balance * 0.74, 2)

def test_export_universal_columns_standard():
    """Verify standard export format contains currency metadata block and universal columns."""
    pdf_path = "tests/clean_pdf_statement.pdf"
    excel_path = "tests/currency_reconciliation_report.xlsx"
    
    try:
        canonical = process_statement(
            pdf_path,
            excel_output_path=excel_path,
            universal_currency="EUR",
            exchange_rate=0.74,
            flat_table=False
        )
        
        assert os.path.exists(excel_path)
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        ws = wb["Reconciliation Ledger"]
        
        # Check metadata currency rows
        found_curr_label = False
        for r in range(1, 15):
            val = ws.cell(row=r, column=1).value
            if val == "Universal Target Currency:":
                assert ws.cell(row=r, column=2).value == "EUR"
                found_curr_label = True
                break
        assert found_curr_label
        
        # Table columns count: Date, Desc, Dr, Cr, Balance + 3 Universal cols + Status, Conf, Notes = 11 cols!
        # Headers should start around row 21 or 24 depending on metadata height
        headers_row = -1
        for r in range(10, 30):
            if ws.cell(row=r, column=1).value == "Date":
                headers_row = r
                break
        assert headers_row != -1
        
        headers = [ws.cell(row=headers_row, column=c).value for c in range(1, 12)]
        assert "Universal Withdrawals (EUR)" in headers
        assert "Universal Deposits (EUR)" in headers
        assert "Universal Balance (EUR)" in headers
        
        # Verify first data row values (Opening Balance)
        data_row = headers_row + 1
        assert ws.cell(row=data_row, column=8).value == 740.00 # Universal balance for 1000.0
        
    finally:
        if os.path.exists(excel_path):
            os.remove(excel_path)
