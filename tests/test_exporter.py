import os
import openpyxl
import pytest
from tests.generate_test_data import create_corrupt_pdf
from soa_guardian.pipeline import process_statement

def test_excel_export_e2e():
    """Verify that E2E pipeline generates a beautifully formatted, mathematically correct Excel report."""
    pdf_path = "tests/corrupt_statement_test.pdf"
    excel_path = "tests/reconciliation_report.xlsx"
    
    # Generate synthetic corrupt statement (OCR error: 12O.50)
    create_corrupt_pdf(pdf_path)
    
    try:
        # Run process_statement with the excel_output_path parameter
        canonical = process_statement(pdf_path, excel_output_path=excel_path)
        
        # Verify output file was created
        assert os.path.exists(excel_path)
        
        # Load workbook with openpyxl (keep formulas visible for auditing)
        wb = openpyxl.load_workbook(excel_path, data_only=False)
        
        # 1. Assert sheet names
        assert "Reconciliation Ledger" in wb.sheetnames
        assert "Audit Logs & Provenance" in wb.sheetnames
        
        ws1 = wb["Reconciliation Ledger"]
        
        # 2. Check title header
        title_val = ws1.cell(row=1, column=1).value
        assert "SoA Guardian" in title_val
        
        # 3. Check metadata values
        assert ws1.cell(row=3, column=1).value == "Vendor / Account Owner:"
        assert ws1.cell(row=3, column=2).value == "SoA Guardian Bank"
        assert ws1.cell(row=7, column=1).value == "Triage Verification status:"
        # Since it had a transcription error that was repaired, the status should be CLEAN / RECONCILED
        assert ws1.cell(row=7, column=2).value == "CLEAN / RECONCILED"
        
        # 4. Check invariants calculations block
        assert "Calculated Statement Ending Balance" in ws1.cell(row=14, column=1).value
        calc_closing_val = ws1.cell(row=14, column=2).value
        assert isinstance(calc_closing_val, float)
        assert calc_closing_val == 3095.25 # Opening 1000 + Credits 2515.75 - Debits 420.50
        
        # 5. Check transactions ledger table
        header_row = 21
        assert ws1.cell(row=header_row, column=1).value == "Date"
        assert ws1.cell(row=header_row, column=2).value == "Description / Particulars"
        assert ws1.cell(row=header_row, column=3).value == "Withdrawals (Dr)"
        
        # Check transaction rows (Row 22: Opening, Row 23: Salary, Row 24: Grocery)
        repaired_row = 24
        assert ws1.cell(row=repaired_row, column=1).value == "03/06/2026"
        assert ws1.cell(row=repaired_row, column=2).value == "Grocery Shop Supermarket Store #55"
        assert ws1.cell(row=repaired_row, column=3).value == 120.5
        assert ws1.cell(row=repaired_row, column=6).value == "AUTO_REPAIRED"
        
        # Check color fills (ARGB formatting in openpyxl)
        green_fill_color = ws1.cell(row=22, column=1).fill.start_color.rgb
        blue_fill_color = ws1.cell(row=repaired_row, column=1).fill.start_color.rgb
        
        # Clean fills should be Soft Green (E2EFDA), Auto-repaired should be Soft Blue (DDEBF7)
        assert green_fill_color in ["00E2EFDA", "FFE2EFDA"]
        assert blue_fill_color in ["00DDEBF7", "FFDDEBF7"]
        
        # 6. Check summary row and mathematical totals formulas
        total_row = 27
        assert ws1.cell(row=total_row, column=1).value == "Ledger Total Summary"
        assert ws1.cell(row=total_row, column=3).value == "=SUM(C22:C26)"
        assert ws1.cell(row=total_row, column=4).value == "=SUM(D22:D26)"
        
        # 7. Check Sheet 2: Audit Logs & Provenance details
        ws2 = wb["Audit Logs & Provenance"]
        assert ws2.cell(row=3, column=1).value == "Auto-Repair Operations"
        assert ws2.cell(row=5, column=1).value == "transactions[2].debit_amount"
        assert ws2.cell(row=5, column=2).value == "12O.50"
        assert ws2.cell(row=5, column=3).value == "120.50"
        assert "OCR confusion repair resolved equations" in ws2.cell(row=5, column=4).value
        
    finally:
        # Clean up files
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        if os.path.exists(excel_path):
            os.remove(excel_path)

def test_flat_table_export_e2e():
    """Verify that E2E pipeline with flat_table=True generates a clean table starting with headers at Row 1."""
    pdf_path = "tests/corrupt_statement_test.pdf"
    excel_path = "tests/flat_reconciliation_report.xlsx"
    
    # Generate synthetic corrupt statement (OCR error: 12O.50)
    create_corrupt_pdf(pdf_path)
    
    try:
        # Run process_statement with flat_table=True
        canonical = process_statement(pdf_path, excel_output_path=excel_path, flat_table=True)
        
        # Verify output file was created
        assert os.path.exists(excel_path)
        
        # Load workbook with openpyxl
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # There should only be the main reconciliation ledger sheet (or audit logs is omitted because it's flat)
        assert "Reconciliation Ledger" in wb.sheetnames
        ws = wb["Reconciliation Ledger"]
        
        # Assert Row 1 is exactly the header row
        headers = [ws.cell(row=1, column=c).value for c in range(1, 6)]
        assert headers == ["Date", "Description", "Withdrawals (Dr)", "Deposits (Cr)", "Balance"]
        
        # Assert Row 2 to 6 contain transaction data directly (no metadata block!)
        row_2 = [ws.cell(row=2, column=c).value for c in range(1, 6)]
        assert row_2 == ["01/06/2026", "Opening Balance", None, None, 1000.0]
        
        row_3 = [ws.cell(row=3, column=c).value for c in range(1, 6)]
        assert row_3 == ["02/06/2026", "Salary Credit", None, 2500.0, 3500.0]
        
        row_4 = [ws.cell(row=4, column=c).value for c in range(1, 6)]
        # Repaired row (12O.50 -> 120.5)
        assert row_4 == ["03/06/2026", "Grocery Shop Supermarket Store #55", 120.5, None, 3379.5]
        
        # Verify no total summaries at the bottom (Row 7 is empty or None)
        assert ws.cell(row=7, column=1).value is None
        
    finally:
        # Clean up files
        if os.path.exists(pdf_path):
            os.remove(pdf_path)
        if os.path.exists(excel_path):
            os.remove(excel_path)
