import os
import json
from typing import List, Dict, Any, Optional, Tuple
from soa_guardian.mapping import infer_header_field

class VendorRegistry:
    """Intelligent database registry to auto-identify vendor profiles in unnamed files."""
    
    def __init__(self, mappings_path: Optional[str] = None):
        self.vendors = {}
        if mappings_path is not None and mappings_path.lower().endswith(".xlsx"):
            if os.path.exists(mappings_path):
                excel_mappings = self._discover_excel_mappings(specific_path=mappings_path)
                if excel_mappings:
                    self.vendors.update(excel_mappings)
            return

        if mappings_path is None:
            # Auto-locate mappings file inside package root
            current_dir = os.path.dirname(os.path.abspath(__file__))
            mappings_path = os.path.join(current_dir, "vendor_mappings.json")
            
        if os.path.exists(mappings_path):
            try:
                with open(mappings_path, "r", encoding="utf-8") as f:
                    self.vendors = json.load(f)
            except Exception as e:
                print(f"Warning: Failed to load vendor_mappings.json: {e}")

        # Auto-discover and parse external Excel mapping sheet if present
        excel_mappings = self._discover_excel_mappings()
        if excel_mappings:
            self.vendors.update(excel_mappings)

    def _discover_excel_mappings(self, specific_path: Optional[str] = None) -> Dict[str, Any]:
        """Scans workspace directories or loads specific path to parse Excel mappings registry."""
        import glob
        import openpyxl
        import re
        
        if specific_path and os.path.exists(specific_path):
            paths = [specific_path]
        else:
            package_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            search_paths = [
                package_root,
                os.path.join(package_root, "tests"),
                os.path.join(package_root, "scratch"),
                os.path.join(package_root, "Internship Problem Statement Docs"),
                os.getcwd(),
                os.path.join(os.getcwd(), "tests"),
                os.path.join(os.getcwd(), "scratch"),
                os.path.join(os.getcwd(), "Internship Problem Statement Docs")
            ]
            
            # Deduplicate paths
            unique_paths = []
            for p in search_paths:
                if p and os.path.exists(p):
                    abs_p = os.path.abspath(p)
                    if abs_p not in unique_paths:
                        unique_paths.append(abs_p)
            
            paths = []
            for base_path in unique_paths:
                paths.extend(glob.glob(os.path.join(base_path, "*.xlsx")))
                
        for path in paths:
            try:
                wb = openpyxl.load_workbook(path, read_only=True, keep_links=False)
                if "Final_mappings" in wb.sheetnames:
                    # Found the mappings file! Let's load the data
                    wb_full = openpyxl.load_workbook(path, data_only=True)
                    ws = wb_full["Final_mappings"]
                    
                    # 1. Read first row headers dynamically
                    first_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                    vendors_idx = -1
                    for idx, h in enumerate(first_row):
                        if h and str(h).strip().lower() == "vendors":
                            vendors_idx = idx
                            break
                    if vendors_idx == -1:
                        vendors_idx = 0
                        
                    # Extract all target column headers
                    target_headers = []
                    target_header_indices = []
                    for idx, h in enumerate(first_row):
                        if idx != vendors_idx and h:
                            target_headers.append(str(h).strip())
                            target_header_indices.append(idx)
                            
                    excel_vendors = {}
                    # 2. Iterate rows and extract dynamically mapped fields
                    for row_idx in range(2, ws.max_row + 1):
                        vendor_val = ws.cell(row=row_idx, column=vendors_idx + 1).value
                        if not vendor_val:
                            continue
                        
                        vendor_name = str(vendor_val).strip()
                        vendor_key = re.sub(r'[^a-z0-9_]', '', re.sub(r'\s+', '_', re.sub(r'\.(xlsx|xls|pdf|csv)', '', vendor_name, flags=re.IGNORECASE)).lower())
                        
                        alias_name = re.sub(r'\.(xlsx|xls|pdf|csv)', '', vendor_name, flags=re.IGNORECASE).strip()
                        
                        mapping = {
                            "official_name": alias_name,
                            "aliases": [alias_name, alias_name.lower(), vendor_name],
                            "columns": {},
                            "output_format_columns": {}
                        }
                        
                        # Populate target mappings dynamically
                        for h_name, h_idx in zip(target_headers, target_header_indices):
                            cell_val = ws.cell(row=row_idx, column=h_idx + 1).value
                            mapping["output_format_columns"][h_name] = str(cell_val).strip() if cell_val else ""
                            
                        # Resolve standard canonical fields using field scoring over the workbook values.
                        def score_value_for_field(value: str, field: str) -> float:
                            if not value:
                                return 0.0
                            text = str(value).strip().lower()
                            if not text:
                                return 0.0
                            
                            words = set(re.sub(r'[^a-z0-9\s]', ' ', text).split())
                            def has_token(token: str) -> bool:
                                if " " in token:
                                    return token in text
                                return token in words

                            if field == "transaction_date":
                                if any(has_token(t) for t in ["transaction date", "txn date", "posting date", "doc date", "value date", "booking date", "payment date", "due date", "date", "posting", "due"]):
                                    return 1.0
                                return 0.0

                            if field == "description":
                                if any(has_token(t) for t in ["document type", "doc type", "description", "particular", "detail", "text", "narrative", "memo", "remarks", "assignment"]):
                                    return 1.0
                                if any(has_token(t) for t in ["reference", "reference key", "invoice", "document no", "invoice no"]):
                                    return 0.35
                                return 0.0

                            if field == "debit_amount":
                                if any(has_token(t) for t in ["debit", "withdrawal", "dr", "amount dr", "payments", "paid", "out"]):
                                    return 1.0
                                if any(has_token(t) for t in ["amount", "amt", "sum", "value"]):
                                    return 0.75
                                return 0.0

                            if field == "credit_amount":
                                if any(has_token(t) for t in ["credit", "deposit", "cr", "amount cr", "receipts", "received", "refund", "in"]):
                                    return 1.0
                                if any(has_token(t) for t in ["amount", "amt", "sum", "value"]):
                                    return 0.75
                                return 0.0

                            if field == "running_balance":
                                if any(has_token(t) for t in ["balance", "cum bal", "cum balance", "running balance", "bal", "closing", "outstanding"]):
                                    return 1.0
                                return 0.0

                            return 0.0

                        field_scores = {}
                        for h_name in target_headers:
                            raw_value = mapping["output_format_columns"].get(h_name, "")
                            if not raw_value:
                                continue
                            for field in ["transaction_date", "description", "debit_amount", "credit_amount", "running_balance"]:
                                score = score_value_for_field(raw_value, field)
                                if score > 0.0:
                                    field_scores.setdefault(h_name, {})[field] = score

                        for h_name, scores in field_scores.items():
                            if not scores:
                                continue
                            best_field = max(scores, key=scores.get)
                            best_score = scores[best_field]
                            if best_score < 0.5:
                                continue
                            if best_field == "debit_amount":
                                mapping["columns"]["debit_amount"] = mapping["output_format_columns"][h_name]
                                mapping["columns"].setdefault("credit_amount", mapping["output_format_columns"][h_name])
                            elif best_field == "credit_amount":
                                mapping["columns"]["credit_amount"] = mapping["output_format_columns"][h_name]
                                mapping["columns"].setdefault("debit_amount", mapping["output_format_columns"][h_name])
                            else:
                                mapping["columns"][best_field] = mapping["output_format_columns"][h_name]

                        if "transaction_date" not in mapping["columns"]:
                            for h_name in target_headers:
                                raw_value = mapping["output_format_columns"].get(h_name, "")
                                if not raw_value:
                                    continue
                                if "date" in str(raw_value).lower() or "posting" in str(raw_value).lower():
                                    mapping["columns"]["transaction_date"] = raw_value
                                    break

                        if "description" not in mapping["columns"]:
                            for h_name in target_headers:
                                raw_value = mapping["output_format_columns"].get(h_name, "")
                                if not raw_value:
                                    continue
                                if any(token in str(raw_value).lower() for token in ["description", "particular", "detail", "text", "narrative", "type", "assignment"]):
                                    mapping["columns"]["description"] = raw_value
                                    break
                        if "debit_amount" not in mapping["columns"] and "credit_amount" not in mapping["columns"]:
                            for h_name in target_headers:
                                raw_value = mapping["output_format_columns"].get(h_name, "")
                                if raw_value and "amount" in str(raw_value).lower():
                                    mapping["columns"]["debit_amount"] = raw_value
                                    mapping["columns"]["credit_amount"] = raw_value
                                    break
                            
                        excel_vendors[vendor_key] = mapping
                    return excel_vendors
            except Exception as e:
                print(f"Warning: Discover Excel mappings error on {path}: {e}")
        return {}

    def identify_vendor(
        self, 
        text_content: str, 
        extracted_headers: List[str],
        similarity_threshold: float = 0.80,
        original_filename: Optional[str] = None
    ) -> Optional[Tuple[str, Dict[str, Any]]]:
        """Identifies vendor by scanning text for aliases, falling back to header structure matching.
        
        Args:
            text_content: Full text layer extracted from statement document.
            extracted_headers: List of raw header columns extracted from table.
            similarity_threshold: Overlap ratio threshold for structural matches (default 80%).
            original_filename: Optional original uploaded filename.
            
        Returns:
            Tuple of (vendor_key, vendor_config) if resolved, else None.
        """
        # Pass 1: Search for aliases in the raw text content or original filename
        import re
        def clean_str(s: str) -> str:
            return re.sub(r"[^a-z0-9]", "", s.lower())
            
        text_norm = clean_str(text_content)
        filename_norm = clean_str(original_filename) if original_filename else ""
        for key, config in self.vendors.items():
            for alias in config.get("aliases", []):
                alias_norm = clean_str(alias)
                if alias_norm:
                    if alias_norm in text_norm or (filename_norm and alias_norm in filename_norm):
                        return key, config
                    
        # Pass 2: Fuzzy header structure overlap with exact-string preference
        best_key = None
        best_score = 0.0
        
        cleaned_extracted = [h.strip() for h in extracted_headers if h.strip()]
        
        if cleaned_extracted:
            for key, config in self.vendors.items():
                vendor_headers = [h.strip() for h in config.get("columns", {}).values()]
                if not vendor_headers:
                    continue

                total_score = 0.0
                for vh in vendor_headers:
                    best_field_score = 0.0
                    for eh in cleaned_extracted:
                        cell_norm = re.sub(r'[^a-z0-9]', '', eh.lower())
                        header_norm = re.sub(r'[^a-z0-9]', '', vh.lower())
                        if not cell_norm or not header_norm:
                            continue
                        if cell_norm == header_norm:
                            best_field_score = 1.0
                            break
                        if cell_norm in header_norm or header_norm in cell_norm:
                            best_field_score = max(best_field_score, 0.85)
                            continue
                        field1 = infer_header_field(eh)
                        field2 = infer_header_field(vh)
                        if field1 and field2 and field1 == field2:
                            best_field_score = max(best_field_score, 0.60)
                    total_score += best_field_score

                score = total_score / len(vendor_headers)
                if score > best_score:
                    best_score = score
                    best_key = key
                    
            if best_score >= similarity_threshold and best_key:
                return best_key, self.vendors[best_key]
                
        return None
