import os
import sys
import re
import csv
import random
from typing import List, Dict, Any, Tuple, Optional
import openpyxl

# List of standard mock business narratives to cycle through
MOCK_DESCRIPTIONS = [
    "Invoice Payment #",
    "Material Procurement Batch #",
    "Logistics & Shipping Services Fee",
    "Consultancy Retainer Service",
    "Equipment Maintenance Lease",
    "Office Supplies Expense",
    "IT Support Ticket Resolution Fee",
    "Marketing Campaign Distribution",
    "Warehouse Storage Space Allocation",
    "Insurance Premium Payment",
    "Tax Remittance",
    "Utility Bill Reimbursement",
    "Vendor Rebate Allocation"
]

def scrub_text(text: str) -> str:
    """Replaces email addresses, phone numbers, customer names, and account numbers with placeholders."""
    if not text:
        return ""
    s = str(text).strip()
    
    # Bypass scrubbing completely if the value is a pure numeric float/int (not PII)
    try:
        float(s.replace(",", "").replace("$", ""))
        return s
    except ValueError:
        pass
    
    # 1. Anonymize Emails
    s = re.sub(r'[\w\.-]+@[\w\.-]+\.\w+', 'contact@mockvendor.com', s)
    
    # 2. Anonymize Phone Numbers (requires space/dash separators to avoid matching plain decimals)
    s = re.sub(r'\+?\d{1,4}[-\s]\d{1,5}(?:[-\s]\d{3,9})?', '+1-555-0199', s)
    
    # 3. Anonymize Account Numbers (e.g., Account Number: 12345)
    s = re.sub(r'(Account\s*(?:No|Number|#)?\s*[:\-]?\s*)\b[A-Za-z0-9\-]+\b', r'\g<1>ACC-XXXXXX', s, flags=re.IGNORECASE)
    
    # 4. Anonymize specific bank/company names in text
    s = re.sub(r'\b(?:Chase|Citi|Wells Fargo|HSBC|Barclays|Apex Bank|Aragen|Sigma|Aldrich)\b', 'Mock-Enterprise', s, flags=re.IGNORECASE)
    
    return s

def identify_header_columns(row: List[str]) -> Dict[str, int]:
    """Identifies the indices of transaction grid column headers using alias matching."""
    mapping = {}
    clean_row = [str(c).strip().lower() for c in row if c is not None]
    
    aliases = {
        "transaction_date": ["date", "txn date", "txndate", "value date", "booking date"],
        "description": ["description", "particulars", "details", "narrative", "remarks", "memo"],
        "debit_amount": ["debit", "withdrawal", "dr", "paid out", "payments"],
        "credit_amount": ["credit", "deposit", "cr", "paid in", "receipts"],
        "running_balance": ["balance", "running balance", "bal", "cumbal", "outstanding"]
    }
    
    for field, terms in aliases.items():
        for i, cell in enumerate(clean_row):
            matched = False
            for t in terms:
                if t == cell:
                    matched = True
                    break
                # Avoid substring matches like 'cr' in 'description' using word boundaries
                if len(t) <= 3:
                    if re.search(rf"\b{re.escape(t)}\b", cell):
                        matched = True
                        break
                else:
                    if t in cell:
                        matched = True
                        break
            if matched:
                mapping[field] = i
                break
    return mapping

def scrub_grid(
    grid: List[List[Any]], 
    randomize_amounts: bool = False
) -> List[List[Any]]:
    """Anonymizes text, and optionally recalculates numeric amounts with valid mathematical balance flow."""
    header_idx = -1
    col_mapping = {}
    
    # Identify header row
    for r_idx, row in enumerate(grid):
        mapping = identify_header_columns(row)
        if "transaction_date" in mapping and "description" in mapping:
            col_mapping = mapping
            header_idx = r_idx
            break
            
    scrubbed_grid = []
    
    # Math accumulators for randomization
    running_balance_accum = 0.0
    opening_balance_found = False
    desc_cycle_idx = 0
    
    for r_idx, row in enumerate(grid):
        row_clean = [str(cell).strip() if cell is not None else "" for cell in row]
        
        # 1. Metadata rows (prior to header)
        if header_idx == -1 or r_idx < header_idx:
            scrubbed_row = [scrub_text(cell) for cell in row_clean]
            
            if randomize_amounts:
                # Rewrite opening balance in metadata cells to keep totals consistent
                for c_idx, cell in enumerate(scrubbed_row):
                    if any(p in cell.lower() for p in ["opening balance", "brought forward"]):
                        # Find numeric amount in the row and replace it
                        for target_idx in range(len(scrubbed_row)):
                            cell_val = scrubbed_row[target_idx]
                            # Look for digits
                            if re.search(r'\d+', cell_val):
                                new_op = float(random.randint(3000, 15000))
                                running_balance_accum = new_op
                                opening_balance_found = True
                                scrubbed_row[target_idx] = re.sub(r'[\d\.,\-]+', f"{new_op:.2f}", cell_val)
                                break
            scrubbed_grid.append(scrubbed_row)
            continue
            
        # 2. Header row itself (keep intact to preserve layout structure)
        if r_idx == header_idx:
            scrubbed_grid.append(row_clean)
            continue
            
        # 3. Transaction/Data rows
        date_col = col_mapping.get("transaction_date", -1)
        desc_col = col_mapping.get("description", -1)
        deb_col = col_mapping.get("debit_amount", -1)
        cred_col = col_mapping.get("credit_amount", -1)
        bal_col = col_mapping.get("running_balance", -1)
        
        scrubbed_row = list(row_clean)
        
        # Check if this row is a transaction row (must contain date elements, not string titles)
        is_transaction = False
        if date_col >= 0 and date_col < len(row_clean) and row_clean[date_col].strip():
            if not any(c.isalpha() for c in row_clean[date_col]):
                is_transaction = True
                
        if not is_transaction:
            # Footer summary row (e.g. "Closing Balance: 3500.00")
            footer_row = [scrub_text(cell) for cell in row_clean]
            if randomize_amounts and opening_balance_found:
                # Rewrite the ending summary balance to match the final calculated total
                for c_idx, cell in enumerate(footer_row):
                    if any(p in cell.lower() for p in ["closing balance", "carried forward"]):
                        for target_idx in range(len(footer_row)):
                            cell_val = footer_row[target_idx]
                            if re.search(r'\d+', cell_val):
                                footer_row[target_idx] = re.sub(r'[\d\.,\-]+', f"{running_balance_accum:.2f}", cell_val)
                                break
            scrubbed_grid.append(footer_row)
            continue
            
        # Transaction Description scrubbing
        if desc_col >= 0 and desc_col < len(scrubbed_row):
            mock_desc = MOCK_DESCRIPTIONS[desc_cycle_idx % len(MOCK_DESCRIPTIONS)]
            txn_num = random.randint(10000, 99999)
            scrubbed_row[desc_col] = f"{mock_desc}{txn_num}"
            desc_cycle_idx += 1
            
        # Scrub other details in metadata columns
        for c_idx in range(len(scrubbed_row)):
            if c_idx not in [date_col, desc_col, deb_col, cred_col, bal_col]:
                scrubbed_row[c_idx] = scrub_text(scrubbed_row[c_idx])
                
        # Optional Amount scrambling with valid running sequence mathematics
        if randomize_amounts:
            orig_deb = row_clean[deb_col].strip() if deb_col >= 0 and deb_col < len(row_clean) else ""
            orig_cred = row_clean[cred_col].strip() if cred_col >= 0 and cred_col < len(row_clean) else ""
            
            new_deb = 0.0
            new_cred = 0.0
            
            if orig_deb and orig_deb not in ["0.0", "0", "0.00"]:
                new_deb = round(random.uniform(5.0, 800.0), 2)
                scrubbed_row[deb_col] = f"{new_deb:.2f}"
            else:
                if deb_col >= 0:
                    scrubbed_row[deb_col] = ""
                    
            if orig_cred and orig_cred not in ["0.0", "0", "0.00"]:
                new_cred = round(random.uniform(50.0, 2000.0), 2)
                scrubbed_row[cred_col] = f"{new_cred:.2f}"
            else:
                if cred_col >= 0:
                    scrubbed_row[cred_col] = ""
                    
            # Recalculate mathematical balance
            if not opening_balance_found and bal_col >= 0:
                try:
                    orig_bal_cleaned = re.sub(r'[^\d\.-]', '', row_clean[bal_col])
                    running_balance_accum = float(orig_bal_cleaned)
                except ValueError:
                    running_balance_accum = 1000.0
                opening_balance_found = True
                
            running_balance_accum += new_cred - new_deb
            if bal_col >= 0 and bal_col < len(scrubbed_row):
                scrubbed_row[bal_col] = f"{running_balance_accum:.2f}"
                
        scrubbed_grid.append(scrubbed_row)
        
    return scrubbed_grid

def scrub_file(input_path: str, output_path: str, randomize_amounts: bool = False):
    """Reads input sheet, scrubs structural values, and saves output."""
    _, ext = os.path.splitext(input_path.lower())
    
    if ext == ".csv":
        with open(input_path, "r", encoding="utf-8", newline="") as f:
            reader = csv.reader(f)
            grid = list(reader)
            
        scrubbed = scrub_grid(grid, randomize_amounts)
        
        with open(output_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(scrubbed)
            
    elif ext in [".xlsx", ".xls"]:
        wb = openpyxl.load_workbook(input_path)
        
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            grid = []
            for row in ws.iter_rows(values_only=True):
                grid.append(list(row))
                
            scrubbed = scrub_grid(grid, randomize_amounts)
            
            # Clear worksheet rows to overwrite cleanly
            ws.calculate_dimension()
            ws.delete_rows(1, ws.max_row + 10)
            
            for r_idx, row in enumerate(scrubbed, 1):
                for c_idx, val in enumerate(row, 1):
                    # Write numbers as floats to keep numeric types intact for testing
                    if r_idx > 1 and str(val).strip():
                        # Simple check for decimals/numbers
                        if re.match(r'^-?\d+\.?\d*$', str(val).strip()):
                            try:
                                ws.cell(row=r_idx, column=c_idx, value=float(val))
                                continue
                            except ValueError:
                                pass
                    ws.cell(row=r_idx, column=c_idx, value=val)
                    
        wb.save(output_path)
    else:
        raise ValueError(f"Unsupported file extension: {ext}. Scrubber supports CSV and XLSX files.")

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python -m soa_guardian.scrubber <input_file> <output_file> [--randomize-amounts]")
        sys.exit(1)
        
    in_file = sys.argv[1]
    out_file = sys.argv[2]
    rand_amt = "--randomize-amounts" in sys.argv
    
    if not os.path.exists(in_file):
        print(f"Error: Input file '{in_file}' does not exist.")
        sys.exit(1)
        
    try:
        scrub_file(in_file, out_file, rand_amt)
        print(f"Success: Anonymized statement saved to '{out_file}' (Randomize amounts: {rand_amt})")
    except Exception as e:
        print(f"Failure: Scrubber failed: {e}")
        sys.exit(1)
