import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from soa_guardian.models import CanonicalStatement

def export_to_excel(
    canonical: CanonicalStatement, 
    output_path: str,
    flat_table: bool = False,
    column_headers: list = None,
    template_path: str = None
):
    """Generates a professionally styled and structured Excel report from CanonicalStatement data."""
    import os
    import shutil
    
    if template_path and os.path.exists(template_path):
        try:
            # Copy template to the output path
            shutil.copy(template_path, output_path)
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Read first row headers
            headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
            
            # Map headers to standard canonical transaction fields
            from soa_guardian.mapping import SemanticMapper
            mapper = SemanticMapper(use_embeddings=False) # Dictionary matching is enough
            
            mapping = {}
            for c_idx, h in enumerate(headers):
                if not h:
                    continue
                h_clean = str(h).strip().lower()
                
                # Match universal fields
                if "universal" in h_clean:
                    if any(k in h_clean for k in ["debit", "withdrawal", "dr"]):
                        mapping["universal_debit"] = c_idx
                    elif any(k in h_clean for k in ["credit", "deposit", "cr"]):
                        mapping["universal_credit"] = c_idx
                    elif any(k in h_clean for k in ["balance", "bal"]):
                        mapping["universal_balance"] = c_idx
                elif "status" in h_clean:
                    mapping["status"] = c_idx
                elif "confidence" in h_clean:
                    mapping["confidence"] = c_idx
                elif any(k in h_clean for k in ["note", "remark", "reason"]):
                    mapping["repair_info_reason"] = c_idx
                else:
                    field, score = mapper.map_header(str(h))
                    if field and field not in mapping:
                        mapping[field] = c_idx
                        
            # Start writing rows from row 2
            start_row = 2
            for tx in canonical.transactions:
                for field, col_idx in mapping.items():
                    cell = ws.cell(row=start_row, column=col_idx + 1)
                    if field == "repair_info_reason":
                        cell.value = tx.repair_info.reason if tx.repair_info else None
                    else:
                        val = getattr(tx, field, None)
                        if val is None:
                            cell.value = None
                        else:
                            cell.value = val
                            
                    # Format number style for numeric fields
                    if field in ["debit_amount", "credit_amount", "running_balance", "universal_debit", "universal_credit", "universal_balance"]:
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = "$#,##0.00"
                start_row += 1
                
            wb.save(output_path)
            return
        except Exception as e:
            print(f"Warning: Template export failed: {e}. Falling back to default generation.")
            
    wb = openpyxl.Workbook()
    # Remove default blank sheet
    default_sheet = wb.active
    if default_sheet is not None:
        wb.remove(default_sheet)
        
    if flat_table:
        ws = wb.create_sheet("Reconciliation Ledger")
        ws.views.sheetView[0].showGridLines = True
        
        # Determine headers
        use_original_indices = False
        use_intended_output_format = False
        if not column_headers:
            if canonical.output_format_columns:
                column_headers = list(canonical.output_format_columns.keys())
                use_intended_output_format = True
            elif canonical.original_headers:
                column_headers = canonical.original_headers
                use_original_indices = True
            else:
                column_headers = [
                    "Date", "Description / Particulars", "Withdrawals (Dr)", "Deposits (Cr)", "Running Balance"
                ]
                if canonical.statement_metadata.exchange_rate != 1.0:
                    univ_cur = canonical.statement_metadata.universal_currency
                    column_headers.extend([
                        f"Universal Withdrawals ({univ_cur})",
                        f"Universal Deposits ({univ_cur})",
                        f"Universal Balance ({univ_cur})"
                    ])
                
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
        white_font_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        normal_font = Font(name="Calibri", size=11)
        thin_side = Side(border_style="thin", color="D3D3D3")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Write headers at Row 1
        for col_idx, h in enumerate(column_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.fill = header_fill
            cell.font = white_font_bold
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_all
            
        ws.row_dimensions[1].height = 25
        
        # Write transactions from Row 2
        for r_idx, tx in enumerate(canonical.transactions, 2):
            if use_intended_output_format:
                # Write in exactly the columns of the dynamic output format
                for col_idx, col_name in enumerate(column_headers, 1):
                    orig_col_name = canonical.output_format_columns.get(col_name, "")
                    cell_val = ""
                    is_numeric = False
                    
                    if not orig_col_name:
                        if "universal" in col_name.lower():
                            if "debit" in col_name.lower() or "dr" in col_name.lower() or "withdrawal" in col_name.lower():
                                cell_val = tx.universal_debit
                                is_numeric = True
                            elif "credit" in col_name.lower() or "cr" in col_name.lower() or "deposit" in col_name.lower():
                                cell_val = tx.universal_credit
                                is_numeric = True
                            elif "balance" in col_name.lower():
                                cell_val = tx.universal_balance
                                is_numeric = True
                            elif "amount" in col_name.lower() or "sum" in col_name.lower() or "val" in col_name.lower():
                                cell_val = tx.universal_debit if tx.universal_debit is not None else tx.universal_credit
                                is_numeric = True
                        else:
                            cell_val = ""
                    elif "date" in col_name.lower():
                        cell_val = tx.transaction_date
                    elif "invoice" in col_name.lower() or "desc" in col_name.lower() or "particular" in col_name.lower():
                        cell_val = tx.description
                    elif "universal" in col_name.lower() and ("amount" in col_name.lower() or "sum" in col_name.lower() or "val" in col_name.lower() or "debit" in col_name.lower() or "credit" in col_name.lower() or "balance" in col_name.lower()):
                        if "debit" in col_name.lower() or "dr" in col_name.lower() or "withdrawal" in col_name.lower():
                            cell_val = tx.universal_debit
                        elif "credit" in col_name.lower() or "cr" in col_name.lower() or "deposit" in col_name.lower():
                            cell_val = tx.universal_credit
                        elif "balance" in col_name.lower():
                            cell_val = tx.universal_balance
                        else:
                            cell_val = tx.universal_debit if tx.universal_debit is not None else tx.universal_credit
                        is_numeric = True
                    elif "amount" in col_name.lower() or "sum" in col_name.lower() or "val" in col_name.lower():
                        cell_val = tx.debit_amount if tx.debit_amount is not None else tx.credit_amount
                        is_numeric = True
                    else:
                        # Fetch custom field using the index of orig_col_name in original statement columns
                        try:
                            orig_idx = canonical.original_headers.index(orig_col_name)
                            cell_val = tx.additional_fields.get(str(orig_idx), "")
                        except ValueError:
                            cell_val = ""
                            
                    cell = ws.cell(row=r_idx, column=col_idx, value=cell_val)
                    cell.font = normal_font
                    cell.border = border_all
                    
                    if "date" in col_name.lower():
                        cell.alignment = Alignment(horizontal="center")
                    elif is_numeric and cell_val is not None:
                        cell.number_format = "$#,##0.00"
            elif use_original_indices:
                # Write cells matching original column sequence indices
                for c_idx in range(len(column_headers)):
                    cell_val = ""
                    is_numeric = False
                    
                    if c_idx == canonical.header_mapping.get("transaction_date"):
                        cell_val = tx.transaction_date
                    elif c_idx == canonical.header_mapping.get("description"):
                        cell_val = tx.description
                    elif c_idx == canonical.header_mapping.get("debit_amount"):
                        cell_val = tx.debit_amount
                        is_numeric = True
                    elif c_idx == canonical.header_mapping.get("credit_amount"):
                        cell_val = tx.credit_amount
                        is_numeric = True
                    elif c_idx == canonical.header_mapping.get("running_balance"):
                        cell_val = tx.running_balance
                        is_numeric = True
                    else:
                        cell_val = tx.additional_fields.get(str(c_idx), "")
                        
                    cell = ws.cell(row=r_idx, column=c_idx + 1, value=cell_val)
                    cell.font = normal_font
                    cell.border = border_all
                    
                    # Formatting
                    if c_idx == canonical.header_mapping.get("transaction_date"):
                        cell.alignment = Alignment(horizontal="center")
                    elif is_numeric and cell_val is not None:
                        cell.number_format = "$#,##0.00"
            else:
                # Standard fallback columns (Date, Desc, Debit, Credit, Balance)
                c_date = ws.cell(row=r_idx, column=1, value=tx.transaction_date)
                c_desc = ws.cell(row=r_idx, column=2, value=tx.description)
                c_deb = ws.cell(row=r_idx, column=3, value=tx.debit_amount)
                c_cred = ws.cell(row=r_idx, column=4, value=tx.credit_amount)
                c_bal = ws.cell(row=r_idx, column=5, value=tx.running_balance)
                
                c_date.alignment = Alignment(horizontal="center")
                if tx.debit_amount is not None:
                    c_deb.number_format = "$#,##0.00"
                if tx.credit_amount is not None:
                    c_cred.number_format = "$#,##0.00"
                c_bal.number_format = "$#,##0.00"
                
                cells_to_format = [c_date, c_desc, c_deb, c_cred, c_bal]
                
                if canonical.statement_metadata.exchange_rate != 1.0:
                    c_udeb = ws.cell(row=r_idx, column=6, value=tx.universal_debit)
                    c_ucred = ws.cell(row=r_idx, column=7, value=tx.universal_credit)
                    c_ubal = ws.cell(row=r_idx, column=8, value=tx.universal_balance)
                    
                    if tx.universal_debit is not None:
                        c_udeb.number_format = "$#,##0.00"
                    if tx.universal_credit is not None:
                        c_ucred.number_format = "$#,##0.00"
                    c_ubal.number_format = "$#,##0.00"
                    cells_to_format.extend([c_udeb, c_ucred, c_ubal])
                
                for cell in cells_to_format:
                    cell.font = normal_font
                    cell.border = border_all
                
        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
        wb.save(output_path)
        wb.close()
        return
        
    # 1. Create Reconciliation & Transactions Ledger sheet
    ws1 = wb.create_sheet("Reconciliation Ledger")
    ws1.views.sheetView[0].showGridLines = True
    
    # Styling Palette
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    white_font_bold = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
    
    # Status color highlights (soft pastel colors)
    status_fills = {
        "clean": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),       # Soft Green
        "auto_repaired": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"), # Soft Blue
        "flagged": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),       # Soft Yellow
        "escalated": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")      # Soft Red
    }
    
    # Borders
    thin_side = Side(border_style="thin", color="D3D3D3")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    double_bottom = Border(
        top=Side(border_style="thin", color="D3D3D3"),
        bottom=Side(border_style="double", color="1F4E78")
    )
    
    row_idx = 1
    
    # --- Title Section ---
    ws1.cell(row=row_idx, column=1, value="SoA Guardian — Verification & Reconciliation Statement").font = title_font
    row_idx += 2
    
    # --- Metadata Block ---
    metadata_fields = [
        ("Vendor / Account Owner:", canonical.statement_metadata.bank_name),
        ("Account Number Reference:", canonical.statement_metadata.account_number),
        ("Statement Period Range:", f"{canonical.statement_metadata.statement_start_date} to {canonical.statement_metadata.statement_end_date}"),
        ("Extraction Confidence Level:", f"{canonical.confidence * 100:.1f}%"),
        ("Triage Verification status:", "REVIEW REQUIRED" if canonical.review_required else "CLEAN / RECONCILED")
    ]
    if canonical.statement_metadata.exchange_rate != 1.0:
        metadata_fields.extend([
            ("Statement Currency:", canonical.statement_metadata.original_currency),
            ("Universal Target Currency:", canonical.statement_metadata.universal_currency),
            ("Applied Exchange Rate:", f"{canonical.statement_metadata.exchange_rate:.4f}")
        ])
        
    for label, val in metadata_fields:
        c1 = ws1.cell(row=row_idx, column=1, value=label)
        c1.font = bold_font
        c2 = ws1.cell(row=row_idx, column=2, value=val)
        c2.font = normal_font
        if label == "Triage Verification status:":
            c2.font = Font(name="Calibri", size=11, bold=True, color="C00000" if canonical.review_required else "375623")
        row_idx += 1
        
    row_idx += 2 # Spacer
    
    # --- Arithmetic Invariant Checks Block ---
    total_debits = sum(tx.debit_amount for tx in canonical.transactions if tx.debit_amount is not None)
    total_credits = sum(tx.credit_amount for tx in canonical.transactions if tx.credit_amount is not None)
    calculated_closing = canonical.opening_balance + total_credits - total_debits
    discrepancy = calculated_closing - canonical.closing_balance
    math_status = "PASS" if abs(discrepancy) <= 0.01 else "FAIL"
    
    ws1.cell(row=row_idx, column=1, value="Arithmetic Invariant Verification Summary").font = bold_font
    row_idx += 1
    
    invariants_data = [
        ("Statement Opening Balance:", canonical.opening_balance),
        ("Total Credits (Deposits):", total_credits),
        ("Total Debits (Withdrawals):", total_debits),
        ("Calculated Statement Ending Balance:", calculated_closing),
        ("Reported Statement Ending Balance:", canonical.closing_balance),
        ("Discrepancy Imbalance (Calculated - Reported):", discrepancy),
        ("Mathematical Equation Check Status:", math_status)
    ]
    if canonical.statement_metadata.exchange_rate != 1.0:
        ex_rate = canonical.statement_metadata.exchange_rate
        univ_cur = canonical.statement_metadata.universal_currency
        invariants_data.extend([
            (f"Universal Opening Balance ({univ_cur}):", canonical.opening_balance * ex_rate),
            (f"Universal Total Credits ({univ_cur}):", total_credits * ex_rate),
            (f"Universal Total Debits ({univ_cur}):", total_debits * ex_rate),
            (f"Universal Calculated Ending ({univ_cur}):", calculated_closing * ex_rate),
        ])
        
    for label, val in invariants_data:
        c1 = ws1.cell(row=row_idx, column=1, value=label)
        c1.font = bold_font
        c2 = ws1.cell(row=row_idx, column=2, value=val)
        
        if isinstance(val, float):
            c2.number_format = "$#,##0.00"
            c2.font = normal_font
        elif label == "Mathematical Equation Check Status:":
            c2.font = Font(name="Calibri", size=11, bold=True, color="375623" if val == "PASS" else "C00000")
        else:
            c2.font = normal_font
        row_idx += 1
        
    row_idx += 2 # Spacer
    
    # --- Transaction ledger Table ---
    ws1.cell(row=row_idx, column=1, value="Transaction Ledger").font = bold_font
    row_idx += 1
    
    table_headers = [
        "Date", "Description / Particulars", "Withdrawals (Dr)", "Deposits (Cr)",
        "Running Balance"
    ]
    has_universal = canonical.statement_metadata.exchange_rate != 1.0
    if has_universal:
        univ_cur = canonical.statement_metadata.universal_currency
        table_headers.extend([
            f"Universal Withdrawals ({univ_cur})",
            f"Universal Deposits ({univ_cur})",
            f"Universal Balance ({univ_cur})"
        ])
    table_headers.extend([
        "Validation Status", "Confidence", "Auto-Repair / Escalation Details"
    ])
    
    # Write header row
    for col_idx, h in enumerate(table_headers, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = white_font_bold
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws1.row_dimensions[row_idx].height = 25
    row_idx += 1
    
    # Write statement transaction rows
    first_tx_row = row_idx
    for tx in canonical.transactions:
        c_date = ws1.cell(row=row_idx, column=1, value=tx.transaction_date)
        c_desc = ws1.cell(row=row_idx, column=2, value=tx.description)
        c_deb = ws1.cell(row=row_idx, column=3, value=tx.debit_amount)
        c_cred = ws1.cell(row=row_idx, column=4, value=tx.credit_amount)
        c_bal = ws1.cell(row=row_idx, column=5, value=tx.running_balance)
        
        curr_col = 6
        cells_to_format = [c_date, c_desc, c_deb, c_cred, c_bal]
        
        if has_universal:
            c_udeb = ws1.cell(row=row_idx, column=6, value=tx.universal_debit)
            c_ucred = ws1.cell(row=row_idx, column=7, value=tx.universal_credit)
            c_ubal = ws1.cell(row=row_idx, column=8, value=tx.universal_balance)
            
            if tx.universal_debit is not None:
                c_udeb.number_format = "$#,##0.00"
            if tx.universal_credit is not None:
                c_ucred.number_format = "$#,##0.00"
            c_ubal.number_format = "$#,##0.00"
            
            cells_to_format.extend([c_udeb, c_ucred, c_ubal])
            curr_col += 3
            
        c_stat = ws1.cell(row=row_idx, column=curr_col, value=tx.status.upper())
        c_conf = ws1.cell(row=row_idx, column=curr_col+1, value=tx.confidence)
        
        # Resolve repair comments for provenance transparency
        notes = ""
        if tx.status == "auto_repaired" and tx.repair_info:
            notes = f"Auto-Repaired: '{tx.repair_info.raw_value}' -> '{tx.repair_info.corrected_value}' (Reason: {tx.repair_info.reason})"
        elif tx.status == "escalated":
            notes = "Escalated: Arithmetic validation failed. Handed to review queue."
            
        c_notes = ws1.cell(row=row_idx, column=curr_col+2, value=notes)
        cells_to_format.extend([c_stat, c_conf, c_notes])
        
        # Apply formatting
        c_date.alignment = Alignment(horizontal="center")
        if tx.debit_amount is not None:
            c_deb.number_format = "$#,##0.00"
        if tx.credit_amount is not None:
            c_cred.number_format = "$#,##0.00"
        c_bal.number_format = "$#,##0.00"
        c_stat.alignment = Alignment(horizontal="center")
        c_conf.number_format = "0.0%"
        c_conf.alignment = Alignment(horizontal="center")
        
        row_fill = status_fills.get(tx.status)
        for cell in cells_to_format:
            cell.font = normal_font
            cell.border = border_all
            if row_fill:
                cell.fill = row_fill
                
        row_idx += 1
    last_tx_row = row_idx - 1
    
    # Add Total summary row at bottom of ledger
    ws1.cell(row=row_idx, column=1, value="Ledger Total Summary").font = bold_font
    ws1.cell(row=row_idx, column=1).border = Border(top=Side(style="thin", color="D3D3D3"))
    
    if len(canonical.transactions) > 0:
        # Sum columns with formulas
        cell_deb_sum = ws1.cell(row=row_idx, column=3, value=f"=SUM(C{first_tx_row}:C{last_tx_row})")
        cell_deb_sum.number_format = "$#,##0.00"
        cell_deb_sum.font = bold_font
        cell_deb_sum.border = double_bottom
        
        cell_cred_sum = ws1.cell(row=row_idx, column=4, value=f"=SUM(D{first_tx_row}:D{last_tx_row})")
        cell_cred_sum.number_format = "$#,##0.00"
        cell_cred_sum.font = bold_font
        cell_cred_sum.border = double_bottom
        
        if has_universal:
            cell_udeb_sum = ws1.cell(row=row_idx, column=6, value=f"=SUM(F{first_tx_row}:F{last_tx_row})")
            cell_udeb_sum.number_format = "$#,##0.00"
            cell_udeb_sum.font = bold_font
            cell_udeb_sum.border = double_bottom
            
            cell_ucred_sum = ws1.cell(row=row_idx, column=7, value=f"=SUM(G{first_tx_row}:G{last_tx_row})")
            cell_ucred_sum.number_format = "$#,##0.00"
            cell_ucred_sum.font = bold_font
            cell_ucred_sum.border = double_bottom
            
    # Apply top borders to all columns
    end_col = 11 if has_universal else 8
    for col_idx in range(2, end_col + 1):
        if col_idx in [3, 4] or (has_universal and col_idx in [6, 7]):
            continue
        cell = ws1.cell(row=row_idx, column=col_idx)
        cell.border = Border(top=Side(style="thin", color="D3D3D3"))
        
    # 2. Create Audit logs sheet (only if anomalies exist)
    if canonical.repair_log or canonical.unparsed_lines:
        ws2 = wb.create_sheet("Audit Logs & Provenance")
        ws2.views.sheetView[0].showGridLines = True
        
        ws2.cell(row=1, column=1, value="Audit logs & Extraction Anomaly Details").font = title_font
        row_idx_2 = 3
        
        if canonical.repair_log:
            ws2.cell(row=row_idx_2, column=1, value="Auto-Repair Operations").font = bold_font
            row_idx_2 += 1
            
            rep_headers = ["Field Location", "Raw Extracted Value", "Repaired Value", "Auto-Resolution reasoning", "Confidence"]
            for col_idx, h in enumerate(rep_headers, 1):
                cell = ws2.cell(row=row_idx_2, column=col_idx, value=h)
                cell.fill = header_fill
                cell.font = white_font_bold
                cell.alignment = Alignment(horizontal="center")
            row_idx_2 += 1
            
            for entry in canonical.repair_log:
                c_f = ws2.cell(row=row_idx_2, column=1, value=entry.get("field"))
                c_raw = ws2.cell(row=row_idx_2, column=2, value=entry.get("raw_value"))
                c_rep = ws2.cell(row=row_idx_2, column=3, value=entry.get("corrected_value"))
                c_reason = ws2.cell(row=row_idx_2, column=4, value=entry.get("reason"))
                c_conf = ws2.cell(row=row_idx_2, column=5, value=entry.get("confidence"))
                c_conf.number_format = "0.0%"
                c_conf.alignment = Alignment(horizontal="center")
                
                for cell in [c_f, c_raw, c_rep, c_reason, c_conf]:
                    cell.font = normal_font
                    cell.border = border_all
                row_idx_2 += 1
                
            row_idx_2 += 2 # Spacer
            
        if canonical.unparsed_lines:
            ws2.cell(row=row_idx_2, column=1, value="Unparsed Lines (Action Required)").font = bold_font
            row_idx_2 += 1
            
            unp_headers = ["Raw Statement Line Text", "Page Number", "Extraction Failure Reason"]
            for col_idx, h in enumerate(unp_headers, 1):
                cell = ws2.cell(row=row_idx_2, column=col_idx, value=h)
                cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid") # Bright red header
                cell.font = white_font_bold
                cell.alignment = Alignment(horizontal="center")
            row_idx_2 += 1
            
            for line in canonical.unparsed_lines:
                c_text = ws2.cell(row=row_idx_2, column=1, value=line.raw_text)
                c_page = ws2.cell(row=row_idx_2, column=2, value=line.source_page)
                c_page.alignment = Alignment(horizontal="center")
                c_reason = ws2.cell(row=row_idx_2, column=3, value=line.reason)
                
                for cell in [c_text, c_page, c_reason]:
                    cell.font = normal_font
                    cell.border = border_all
                    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
                row_idx_2 += 1
                
        # Auto-adjust column widths for Sheet 2
        for col in ws2.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws2.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    # Auto-adjust column widths for Sheet 1
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        notes_col = 11 if has_universal else 8
        if col[0].column == notes_col:
            ws1.column_dimensions[col_letter].width = 45
            for cell in col:
                cell.alignment = Alignment(wrap_text=True)
        else:
            ws1.column_dimensions[col_letter].width = max(max_len + 3, 12)
            
    wb.save(output_path)
    wb.close()
